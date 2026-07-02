"""
Excel Bid Sheet Exporter

This module creates professional Excel bid sheets from Bid objects.
The output matches contractor templates with proper formatting.
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from src.models.bid import Bid
from datetime import date


class ExcelBidExporter:
    """
    Exports Bid objects to formatted Excel spreadsheets.

    Creates a professional bid sheet with:
    - Sheet 1: Bill of Materials (itemized list)
    - Sheet 2: Cost Summary (totals by section)
    - Sheet 3: Final Bid (with markup)
    """

    def __init__(self):
        """Initialize exporter with standard formatting styles."""
        # Define reusable styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.section_font = Font(name='Arial', size=11, bold=True)
        self.section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_bid(self, bid: Bid, output_path: Path, workers: list = None) -> Path:
        """
        Export a Bid to Excel file.

        Args:
            bid: The Bid object to export
            output_path: Where to save the Excel file
            workers: Optional list of worker dicts with 'name', 'wage_per_hour', 'hours'

        Returns:
            Path to the created file
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create three sheets
        self._create_bill_of_materials_sheet(wb, bid)
        self._create_cost_summary_sheet(wb, bid)
        self._create_final_bid_sheet(wb, bid, workers)

        # Save workbook
        wb.save(output_path)
        return output_path

    def _create_bill_of_materials_sheet(self, wb: Workbook, bid: Bid):
        """Create detailed Bill of Materials sheet."""
        ws = wb.create_sheet("Bill of Materials")

        # Title
        ws['A1'] = f"Lightning Protection Bid - {bid.project_name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Date: {date.today().strftime('%m/%d/%Y')}"

        # Headers
        headers = ['Section', 'Item Code', 'Description', 'Material', 'Qty', 'Unit', 'Unit Price', 'Material Cost', 'Labor Cost', 'Total', 'Notes']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Data rows
        current_row = 5
        for section in bid.sections:
            # Section header row
            ws.cell(row=current_row, column=1, value=section.name)
            ws.cell(row=current_row, column=1).font = self.section_font
            ws.cell(row=current_row, column=1).fill = self.section_fill
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = self.section_fill
            current_row += 1

            # Line items
            for item in section.line_items:
                ws.cell(row=current_row, column=1, value="")  # Empty section column
                ws.cell(row=current_row, column=2, value=item.price_item.code)
                ws.cell(row=current_row, column=3, value=item.price_item.name)
                ws.cell(row=current_row, column=4, value=item.price_item.material_type or "")
                ws.cell(row=current_row, column=5, value=item.quantity)
                ws.cell(row=current_row, column=6, value=item.price_item.unit or "ea")
                ws.cell(row=current_row, column=7, value=item.price_item.unit_price)
                ws.cell(row=current_row, column=8, value=item.material_cost)
                ws.cell(row=current_row, column=9, value=item.labor_cost or 0)
                ws.cell(row=current_row, column=10, value=item.material_cost + (item.labor_cost or 0))
                ws.cell(row=current_row, column=11, value=item.reason or "")

                # Format currency cells
                for col in [7, 8, 9, 10]:
                    ws.cell(row=current_row, column=col).number_format = '$#,##0.00'

                current_row += 1

            current_row += 1  # Blank row between sections

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 40

    def _create_cost_summary_sheet(self, wb: Workbook, bid: Bid):
        """Create Cost Summary sheet with section totals."""
        ws = wb.create_sheet("Cost Summary")

        # Title
        ws['A1'] = "Cost Summary by Section"
        ws['A1'].font = Font(size=14, bold=True)

        # Headers
        headers = ['Section', 'Material Cost', 'Labor Cost', 'Section Total']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Section totals
        current_row = 4
        for section in bid.sections:
            ws.cell(row=current_row, column=1, value=section.name)
            ws.cell(row=current_row, column=2, value=section.total_material)
            ws.cell(row=current_row, column=3, value=section.total_labor)
            ws.cell(row=current_row, column=4, value=section.section_total)

            # Format currency
            for col in [2, 3, 4]:
                ws.cell(row=current_row, column=col).number_format = '$#,##0.00'

            current_row += 1

        # Totals row
        current_row += 1
        ws.cell(row=current_row, column=1, value="TOTAL (Material + Labor)")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=bid.subtotal_material)
        ws.cell(row=current_row, column=3, value=bid.subtotal_labor)
        ws.cell(row=current_row, column=4, value=bid.subtotal_material + bid.subtotal_labor)

        for col in [2, 3, 4]:
            cell = ws.cell(row=current_row, column=col)
            cell.number_format = '$#,##0.00'
            cell.font = Font(bold=True)
            cell.fill = self.section_fill

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_final_bid_sheet(self, wb: Workbook, bid: Bid, workers: list = None):
        """Create Final Bid sheet with markup and totals."""
        ws = wb.create_sheet("Final Bid")

        # Title
        ws['A1'] = f"Final Bid - {bid.project_name}"
        ws['A1'].font = Font(size=14, bold=True)

        # Material costs section
        row = 3
        ws.cell(row=row, column=1, value="MATERIALS:")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row=row, column=1, value="  Material Cost:")
        ws.cell(row=row, column=2, value=bid.subtotal_material)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1
        
        # Show shipping if > 0
        if bid.shipping_amount > 0:
            ws.cell(row=row, column=1, value="  Shipping:")
            ws.cell(row=row, column=2, value=bid.shipping_amount)
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            row += 1
        
        # Subtotal (Material + Shipping)
        ws.cell(row=row, column=1, value="  Subtotal (Material + Shipping):")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=bid.material_with_shipping)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1
        
        # Use tax if > 0
        if bid.use_tax_pct > 0:
            ws.cell(row=row, column=1, value=f"  Use Tax ({bid.use_tax_pct}%):")
            ws.cell(row=row, column=2, value=bid.material_tax)
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            row += 1
        
        # Total material with tax
        ws.cell(row=row, column=1, value="  Total Material with Tax:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=bid.material_total_with_tax)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).fill = self.section_fill
        row += 2

        # Labor breakdown - show individual workers if provided
        if workers and len(workers) > 0:
            ws.cell(row=row, column=1, value="LABOR:")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            
            # Worker details header
            ws.cell(row=row, column=1, value="  Worker")
            ws.cell(row=row, column=2, value="Wage/Hour")
            ws.cell(row=row, column=3, value="Hours")
            ws.cell(row=row, column=4, value="Total Cost")
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, size=10)
                cell.fill = self.section_fill
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Each worker
            total_labor = 0
            for worker in workers:
                worker_name = worker.get("name", "Worker")
                wage = worker.get("wage_per_hour", 0)
                hours = worker.get("hours", 0)
                worker_cost = wage * hours
                total_labor += worker_cost
                
                ws.cell(row=row, column=1, value=f"  {worker_name}")
                ws.cell(row=row, column=2, value=wage)
                ws.cell(row=row, column=2).number_format = '$#,##0.00'
                ws.cell(row=row, column=3, value=hours)
                ws.cell(row=row, column=3).number_format = '0.0'
                ws.cell(row=row, column=4, value=worker_cost)
                ws.cell(row=row, column=4).number_format = '$#,##0.00'
                row += 1
            
            # Total workers row
            total_hours = sum(w.get("hours", 0) for w in workers)
            ws.cell(row=row, column=1, value="  TOTAL LABOR:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{len(workers)} worker(s)")
            ws.cell(row=row, column=3, value=total_hours)
            ws.cell(row=row, column=3).number_format = '0.0'
            ws.cell(row=row, column=3).font = Font(bold=True)
            ws.cell(row=row, column=4, value=total_labor)
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=4).fill = self.section_fill
            row += 1
        else:
            # Fallback if no worker data provided
            ws.cell(row=row, column=1, value="LABOR:")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="  Total Labor:")
            ws.cell(row=row, column=2, value=bid.subtotal_labor)
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="SUBTOTAL (Material + Labor + Shipping + Tax):")
        ws.cell(row=row, column=2, value=bid.subtotal)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        row += 2

        # Markup
        ws.cell(row=row, column=1, value=f"Material Markup ({bid.material_markup_pct}%):")
        mat_markup = bid.subtotal_material * (bid.material_markup_pct / 100)
        ws.cell(row=row, column=2, value=mat_markup)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1

        ws.cell(row=row, column=1, value=f"Labor Markup ({bid.labor_markup_pct}%):")
        lab_markup = bid.subtotal_labor * (bid.labor_markup_pct / 100)
        ws.cell(row=row, column=2, value=lab_markup)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1

        ws.cell(row=row, column=1, value="Total with Markup:")
        ws.cell(row=row, column=2, value=bid.total_with_markup)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 2

        # Overhead & Profit (calculated on original subtotal, not marked-up amount)
        ws.cell(row=row, column=1, value=f"Overhead ({bid.overhead_pct}%):")
        overhead = bid.subtotal * (bid.overhead_pct / 100)
        ws.cell(row=row, column=2, value=overhead)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1

        ws.cell(row=row, column=1, value=f"Profit ({bid.profit_pct}%):")
        profit = bid.subtotal * (bid.profit_pct / 100)
        ws.cell(row=row, column=2, value=profit)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1
        
        # Additional flat costs (if any)
        if bid.commission_amount > 0:
            ws.cell(row=row, column=1, value="Commission:")
            ws.cell(row=row, column=2, value=bid.commission_amount)
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            row += 1
        
        if bid.tools_rental_amount > 0:
            # Show label with type ($ or %)
            if bid.tools_rental_type == "%":
                label = f"Tools & Rental ({bid.tools_rental_amount}%):"
            else:
                label = "Tools & Rental:"
            
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=bid.tools_rental_cost)
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            row += 1
        
        row += 1

        # FINAL BID
        ws.cell(row=row, column=1, value="FINAL BID AMOUNT:")
        ws.cell(row=row, column=2, value=bid.final_bid_amount)
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15

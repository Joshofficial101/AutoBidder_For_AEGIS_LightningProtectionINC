"""
Enhanced Excel Loader for Lightning Protection Pricing Sheets

This module handles complex Excel formats including:
- Multiple sheets
- Variable header row positions
- Merged cells
- Different column name variations
- Missing or empty rows
"""

from pathlib import Path
import pandas as pd
import openpyxl
from typing import List, Optional, Dict, Tuple
from src.models.items import PriceItem


def _find_header_row(ws, max_rows_to_check: int = 20) -> Optional[int]:
    """
    Find the row that contains column headers.
    
    Looks for common header keywords in the first few rows.
    Returns 0-indexed row number (for pandas) or None if not found.
    """
    header_keywords = [
        "code", "item", "part", "description", "name", "price", "cost",
        "unit", "uom", "labor", "material", "type", "category"
    ]
    
    for row_idx in range(min(max_rows_to_check, ws.max_row)):
        row_values = [str(cell.value or "").lower().strip() for cell in ws[row_idx + 1]]
        # Check if this row contains multiple header keywords
        matches = sum(1 for val in row_values for keyword in header_keywords if keyword in val)
        if matches >= 3:  # Found at least 3 header keywords
            return row_idx
    
    return 0  # Default to first row


def _normalize_column_name(name: str) -> str:
    """Normalize column names for matching."""
    if pd.isna(name) or name is None:
        return ""
    return str(name).strip().lower().replace("_", " ").replace("-", " ")


def _map_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    """
    Map DataFrame columns to our expected fields.
    
    Returns a dictionary mapping field names to column names.
    """
    col_map = {}
    normalized_cols = {_normalize_column_name(col): col for col in df.columns}
    
    # Code/Part Number
    code_patterns = ["code", "item code", "part", "part #", "part number", "part no", "sku", "item number", "no", "no.", "number", "#"]
    col_map["code"] = next((normalized_cols.get(pat) for pat in code_patterns if pat in normalized_cols), None)
    
    # Name/Description
    name_patterns = ["name", "description", "desc", "item", "material name", "product", "item name"]
    col_map["name"] = next((normalized_cols.get(pat) for pat in name_patterns if pat in normalized_cols), None)
    
    # Material Type (prefer actual MATERIAL column over category)
    material_patterns = ["material", "material type"]
    col_map["material_type"] = next((normalized_cols.get(pat) for pat in material_patterns if pat in normalized_cols), None)
    
    # If no material column, fall back to category
    if not col_map["material_type"]:
        type_patterns = ["type", "category", "cat", "class"]
        col_map["material_type"] = next((normalized_cols.get(pat) for pat in type_patterns if pat in normalized_cols), None)
    
    # Unit
    unit_patterns = ["unit", "uom", "units", "unit of measure", "measure"]
    col_map["unit"] = next((normalized_cols.get(pat) for pat in unit_patterns if pat in normalized_cols), None)
    
    # Unit Price (prioritize LIST PRICE for ERICO files)
    price_patterns = ["list price", "price", "unit price", "unit_cost", "cost", "unit cost", "price each", "each"]
    col_map["unit_price"] = next((normalized_cols.get(pat) for pat in price_patterns if pat in normalized_cols), None)
    
    # Labor Rate
    labor_patterns = ["labor", "labor rate", "labor_cost", "labor cost", "install", "installation", "labor each"]
    col_map["labor_rate"] = next((normalized_cols.get(pat) for pat in labor_patterns if pat in normalized_cols), None)
    
    return col_map


def _parse_price(value: any) -> Optional[float]:
    """Parse a price value from various formats."""
    if pd.isna(value) or value is None:
        return None
    
    try:
        # Convert to string and clean
        price_str = str(value).strip()
        if not price_str or price_str.lower() in ["nan", "none", "n/a", ""]:
            return None
        
        # Remove currency symbols and commas
        price_str = price_str.replace("$", "").replace(",", "").replace(" ", "")
        
        # Handle parentheses (negative values)
        if price_str.startswith("(") and price_str.endswith(")"):
            price_str = "-" + price_str[1:-1]
        
        return float(price_str)
    except (ValueError, AttributeError):
        return None


def _load_sheet(path: Path, sheet_name: Optional[str] = None, header_row: Optional[int] = None) -> pd.DataFrame:
    """
    Load a specific sheet from Excel file.
    
    Args:
        path: Path to Excel file
        sheet_name: Name of sheet to load (None = first sheet)
        header_row: Row index to use as header (None = auto-detect)
    
    Returns:
        DataFrame with data
    """
    # Use openpyxl to find header row if not specified
    if header_row is None:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        header_row = _find_header_row(ws)
        wb.close()
    
    # Load with pandas
    try:
        df = pd.read_excel(
            path,
            sheet_name=sheet_name,
            header=header_row,
            dtype=str,  # Read as strings first to avoid type issues
            na_values=["", "N/A", "n/a", "NULL", "null", "None"]
        )
        
        # Clean column names
        df.columns = [_normalize_column_name(col) for col in df.columns]
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        return df
    except Exception as e:
        raise ValueError(f"Error loading sheet '{sheet_name or 'default'}': {str(e)}")


def load_pricing_from_excel(path: Path, sheet_name: Optional[str] = None, 
                           try_all_sheets: bool = True) -> List[PriceItem]:
    """
    Load pricing items from Excel file.
    
    Enhanced version that handles:
    - Multiple sheets (tries all if sheet_name not specified)
    - Auto-detection of header rows
    - Merged cells (handled by openpyxl/pandas)
    - Various column name formats
    - Missing or empty rows
    
    Args:
        path: Path to Excel file
        sheet_name: Specific sheet name to load (None = try all sheets)
        try_all_sheets: If True and sheet_name is None, try all sheets until one works
    
    Returns:
        List of PriceItem objects
    
    Raises:
        ValueError: If file cannot be read or no valid data found
    """
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    
    all_items: List[PriceItem] = []
    sheets_tried = []
    
    # Get list of sheets to try
    if sheet_name:
        sheets_to_try = [sheet_name]
    elif try_all_sheets:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheets_to_try = wb.sheetnames
            wb.close()
        except Exception:
            sheets_to_try = [None]  # Try default sheet
    else:
        sheets_to_try = [None]
    
    # Try each sheet
    for current_sheet in sheets_to_try:
        try:
            df = _load_sheet(path, sheet_name=current_sheet)
            sheets_tried.append(current_sheet or "default")
            
            if df.empty:
                print(f"DEBUG: Sheet '{current_sheet or 'default'}' is empty")
                continue
            
            print(f"DEBUG: Sheet '{current_sheet or 'default'}' has columns: {list(df.columns)}")
            
            # Map columns
            col_map = _map_columns(df)
            print(f"DEBUG: Column mapping: {col_map}")
            
            # Check for required columns (code is optional if we have name and price)
            # We need at least: (name OR code) AND unit_price
            has_name_or_code = col_map["name"] is not None or col_map["code"] is not None
            has_price = col_map["unit_price"] is not None
            
            if not has_price:
                print(f"DEBUG: Missing price column, trying to find by pattern...")
                # Try to find price column by looking for numeric columns
                for col in df.columns:
                    try:
                        # Try to parse first few non-empty values as numbers
                        sample_values = df[col].dropna().head(10)
                        numeric_count = 0
                        for val in sample_values:
                            price = _parse_price(val)
                            if price is not None and price > 0 and price < 100000:  # Reasonable price range
                                numeric_count += 1
                        
                        if numeric_count >= 3:  # At least 3 valid prices found
                            col_map["unit_price"] = col
                            print(f"DEBUG: Found price column by pattern: {col}")
                            has_price = True
                            break
                    except:
                        pass
            
            if not has_name_or_code:
                print(f"DEBUG: Missing name/code column")
                continue
            
            if not has_price:
                print(f"DEBUG: Missing price column, trying next sheet")
                continue
            
            print(f"DEBUG: All required columns found! Name: {col_map['name']}, Code: {col_map['code']}, Price: {col_map['unit_price']}")
            
            # Parse rows
            items = []
            rows_processed = 0
            for _, row in df.iterrows():
                rows_processed += 1
                
                # Parse price first (required) - skip if no valid price
                unit_price = None
                if col_map["unit_price"]:
                    unit_price = _parse_price(row.get(col_map["unit_price"]))
                
                if unit_price is None or unit_price <= 0:
                    continue  # Skip rows without valid price
                
                # Extract code and name fields
                code = ""
                if col_map["code"]:
                    code_val = row.get(col_map["code"], "")
                    if not pd.isna(code_val):
                        code = str(code_val).strip()
                
                name = ""
                if col_map["name"]:
                    name_val = row.get(col_map["name"], "")
                    if not pd.isna(name_val):
                        name = str(name_val).strip()
                
                # If no code but we have name, generate code from row number
                if not code and name:
                    code = f"ITEM-{len(items)+1}"
                
                # Skip only if BOTH code and name are empty (after code generation attempt)
                if not code and not name:
                    continue  # Skip rows with no identifier
                
                # Parse labor rate (optional)
                labor_rate = None
                if col_map["labor_rate"]:
                    labor_rate = _parse_price(row.get(col_map["labor_rate"]))
                
                material_type = None
                if col_map["material_type"]:
                    val = row.get(col_map["material_type"])
                    if not pd.isna(val):
                        material_type = str(val).strip()
                
                unit = None
                if col_map["unit"]:
                    val = row.get(col_map["unit"])
                    if not pd.isna(val):
                        unit = str(val).strip()
                
                # Create PriceItem
                items.append(PriceItem(
                    code=code or f"ITEM-{len(items)+1}",
                    name=name or "Unnamed Item",
                    material_type=material_type if material_type else None,
                    unit=unit if unit else None,
                    unit_price=unit_price,
                    labor_rate=labor_rate
                ))
            
            print(f"DEBUG: Processed {rows_processed} rows, found {len(items)} valid items")
            
            if items:
                all_items.extend(items)
                print(f"DEBUG: Successfully loaded {len(items)} items from sheet '{current_sheet or 'default'}'")
                # If we found items and sheet_name was specified, return immediately
                if sheet_name:
                    return all_items
        
        except Exception as e:
            # Log error but continue trying other sheets
            print(f"DEBUG: Error processing sheet '{current_sheet or 'default'}': {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    
    if not all_items:
        raise ValueError(
            f"Could not extract pricing data from {path.name}.\n"
            f"Tried sheets: {', '.join(sheets_tried)}\n"
            f"Please check that the Excel file has columns for: Code/Part, Name/Description, and Price."
        )
    
    return all_items
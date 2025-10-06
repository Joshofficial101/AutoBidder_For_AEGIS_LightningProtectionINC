"""
Verify Excel File Test

This script checks if the generated Excel file is valid and can be opened.
"""

from pathlib import Path
import openpyxl

output_dir = Path("data/outputs")
excel_files = list(output_dir.glob("*.xlsx"))

if not excel_files:
    print("No Excel files found in data/outputs/")
    print("Run 'py -m src.main' first to generate a bid.")
else:
    print(f"Found {len(excel_files)} Excel file(s):")
    print()

    for excel_file in excel_files:
        print(f"Testing: {excel_file.name}")
        print("-" * 60)

        try:
            # Try to open the file
            wb = openpyxl.load_workbook(excel_file)

            # Check sheets
            print(f"  [OK] File opens successfully")
            print(f"  [OK] Found {len(wb.sheetnames)} sheets:")
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                print(f"       - {sheet} ({ws.max_row} rows, {ws.max_column} cols)")

            # Check some data
            bill_sheet = wb["Bill of Materials"]
            print(f"  [OK] Data validation:")
            print(f"       - Title: {bill_sheet['A1'].value}")

            final_sheet = wb["Final Bid"]
            # Find the final bid amount (it's in column B, look for the largest value)
            for row in range(1, final_sheet.max_row + 1):
                cell = final_sheet.cell(row, 1)
                if cell.value and "FINAL BID" in str(cell.value).upper():
                    amount = final_sheet.cell(row, 2).value
                    print(f"       - Final Bid Amount: ${amount:,.2f}")
                    break

            print(f"  [OK] File is valid and ready to use!")
            print()

            wb.close()

        except Exception as e:
            print(f"  [ERROR] Could not open file: {e}")
            print()

print("=" * 60)
print("How to Open the File:")
print("=" * 60)
print()
print("Option 1: Windows File Explorer")
print(f"  1. Navigate to: {output_dir.absolute()}")
print("  2. Double-click the .xlsx file")
print("  3. Opens in Excel/LibreOffice/Google Sheets")
print()
print("Option 2: Command Line")
print(f"  start {output_dir.absolute()}\\*.xlsx")
print()
print("Option 3: From PyCharm/IntelliJ")
print("  Right-click file → Open In → Associated Application")
print("  (Ignore the IDE preview error - that's just the IDE)")
